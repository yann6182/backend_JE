#!/usr/bin/env python3
"""
Analyse détaillée d'un fichier DPGF spécifique.
Permet de comprendre pourquoi un fichier identifié comme DPGF ne produit pas de données.
"""

import os
import sys
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.utils import get_column_letter

# Ajouter le répertoire scripts au path
scripts_dir = Path(__file__).parent / "scripts"
if scripts_dir.exists():
    sys.path.insert(0, str(scripts_dir))

class DPGFFileAnalyzer:
    """Analyseur détaillé d'un fichier DPGF"""
    
    def __init__(self, file_path: str):
        """
        Initialise l'analyseur
        
        Args:
            file_path: Chemin vers le fichier DPGF à analyser
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Fichier non trouvé: {file_path}")
        
        self.workbook = None
        self.sheets_info = {}
        self.analysis_results = {}
    
    def load_file(self):
        """Charge le fichier Excel"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            print(f"✅ Fichier chargé: {self.file_path.name}")
            print(f"📊 Feuilles trouvées: {len(self.workbook.sheetnames)}")
            
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                self.sheets_info[sheet_name] = {
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'has_data': sheet.max_row > 1 or sheet.max_column > 1
                }
                print(f"   • {sheet_name}: {sheet.max_row} lignes × {sheet.max_column} colonnes")
            
        except Exception as e:
            raise Exception(f"Erreur lors du chargement: {e}")
    
    def detect_dpgf_patterns(self, sheet_name: str = None) -> Dict:
        """
        Détecte les patterns DPGF dans une feuille
        
        Args:
            sheet_name: Nom de la feuille (première feuille si None)
            
        Returns:
            Dict: Patterns détectés
        """
        if not self.workbook:
            self.load_file()
        
        if sheet_name is None:
            sheet_name = self.workbook.sheetnames[0]
        
        sheet = self.workbook[sheet_name]
        
        patterns = {
            'lot_indicators': [],
            'section_indicators': [],
            'price_columns': [],
            'unit_columns': [],
            'quantity_columns': [],
            'designation_columns': [],
            'total_indicators': [],
            'structure_analysis': {
                'has_headers': False,
                'header_row': None,
                'data_starts_at': None,
                'potential_lot_rows': [],
                'potential_section_rows': [],
                'potential_element_rows': []
            }
        }
        
        # Mots-clés DPGF
        lot_keywords = [
            'lot', 'tranche', 'phase', 'ouvrage', 'corps', 'état',
            'travaux', 'marché', 'sous-traitance'
        ]
        
        section_keywords = [
            'chapitre', 'sous-chapitre', 'section', 'sous-section',
            'titre', 'sous-titre', 'poste', 'sous-poste'
        ]
        
        price_keywords = [
            'prix', 'montant', 'coût', 'tarif', 'pu', 'p.u',
            'unitaire', 'total', 'ht', 'ttc'
        ]
        
        unit_keywords = [
            'unité', 'u', 'unite', 'mesure', 'ml', 'm²', 'm3',
            'kg', 'tonnes', 'forfait', 'ens', 'nb', 'pcs'
        ]
        
        quantity_keywords = [
            'quantité', 'qte', 'quantite', 'nombre', 'nb',
            'métré', 'metre', 'volume', 'surface'
        ]
        
        # Analyser chaque ligne
        for row_idx in range(1, min(sheet.max_row + 1, 200)):  # Limiter à 200 lignes pour la performance
            row_data = []
            has_content = False
            
            for col_idx in range(1, min(sheet.max_column + 1, 20)):  # Limiter à 20 colonnes
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                if value is not None:
                    has_content = True
                    row_data.append(str(value).strip())
                else:
                    row_data.append("")
            
            if not has_content:
                continue
            
            # Analyser le contenu de la ligne
            row_text = " ".join(row_data).lower()
            
            # Détecter les indicateurs de lot
            for keyword in lot_keywords:
                if keyword in row_text and len(row_text) < 200:  # Éviter les lignes trop longues
                    patterns['lot_indicators'].append({
                        'row': row_idx,
                        'keyword': keyword,
                        'content': row_text[:100],
                        'columns_with_data': [i for i, val in enumerate(row_data) if val]
                    })
                    patterns['structure_analysis']['potential_lot_rows'].append(row_idx)
                    break
            
            # Détecter les indicateurs de section
            for keyword in section_keywords:
                if keyword in row_text and len(row_text) < 200:
                    patterns['section_indicators'].append({
                        'row': row_idx,
                        'keyword': keyword,
                        'content': row_text[:100],
                        'columns_with_data': [i for i, val in enumerate(row_data) if val]
                    })
                    patterns['structure_analysis']['potential_section_rows'].append(row_idx)
                    break
            
            # Détecter les colonnes de prix/unité/quantité dans les en-têtes
            if row_idx <= 10:  # Chercher dans les 10 premières lignes
                for col_idx, cell_value in enumerate(row_data):
                    if cell_value:
                        cell_lower = cell_value.lower()
                        
                        # Colonnes de prix
                        if any(keyword in cell_lower for keyword in price_keywords):
                            patterns['price_columns'].append({
                                'column': col_idx + 1,
                                'header': cell_value,
                                'row': row_idx
                            })
                        
                        # Colonnes d'unité
                        if any(keyword in cell_lower for keyword in unit_keywords):
                            patterns['unit_columns'].append({
                                'column': col_idx + 1,
                                'header': cell_value,
                                'row': row_idx
                            })
                        
                        # Colonnes de quantité
                        if any(keyword in cell_lower for keyword in quantity_keywords):
                            patterns['quantity_columns'].append({
                                'column': col_idx + 1,
                                'header': cell_value,
                                'row': row_idx
                            })
            
            # Détecter les éléments potentiels (lignes avec prix et unité)
            numeric_values = []
            for cell_value in row_data:
                try:
                    if cell_value and cell_value.replace(',', '.').replace(' ', '').replace('€', '').isdigit():
                        numeric_values.append(float(cell_value.replace(',', '.').replace(' ', '').replace('€', '')))
                except:
                    pass
            
            if len(numeric_values) >= 2 and len(row_text) > 10:  # Au moins 2 valeurs numériques et du texte
                patterns['structure_analysis']['potential_element_rows'].append({
                    'row': row_idx,
                    'numeric_count': len(numeric_values),
                    'content_length': len(row_text),
                    'content_preview': row_text[:50]
                })
        
        # Analyser la structure générale
        if patterns['price_columns'] or patterns['unit_columns']:
            patterns['structure_analysis']['has_headers'] = True
            header_rows = []
            for col_info in patterns['price_columns'] + patterns['unit_columns']:
                header_rows.append(col_info['row'])
            if header_rows:
                patterns['structure_analysis']['header_row'] = min(header_rows)
                patterns['structure_analysis']['data_starts_at'] = min(header_rows) + 1
        
        return patterns
    
    def analyze_file_structure(self) -> Dict:
        """Analyse complète de la structure du fichier"""
        if not self.workbook:
            self.load_file()
        
        analysis = {
            'file_info': {
                'name': self.file_path.name,
                'size': self.file_path.stat().st_size,
                'sheets_count': len(self.workbook.sheetnames),
                'sheets': list(self.workbook.sheetnames)
            },
            'dpgf_compatibility': {
                'score': 0.0,
                'reasons': [],
                'missing_elements': [],
                'detected_elements': []
            },
            'sheets_analysis': {}
        }
        
        total_score = 0
        max_score = 100
        
        # Analyser chaque feuille
        for sheet_name in self.workbook.sheetnames:
            patterns = self.detect_dpgf_patterns(sheet_name)
            analysis['sheets_analysis'][sheet_name] = patterns
            
            # Calculer le score de compatibilité DPGF
            sheet_score = 0
            
            # Points pour les indicateurs de structure
            if patterns['lot_indicators']:
                sheet_score += 25
                analysis['dpgf_compatibility']['detected_elements'].append(f"Lots détectés ({len(patterns['lot_indicators'])})")
            else:
                analysis['dpgf_compatibility']['missing_elements'].append("Aucun lot détecté")
            
            if patterns['section_indicators']:
                sheet_score += 20
                analysis['dpgf_compatibility']['detected_elements'].append(f"Sections détectées ({len(patterns['section_indicators'])})")
            else:
                analysis['dpgf_compatibility']['missing_elements'].append("Aucune section détectée")
            
            if patterns['price_columns']:
                sheet_score += 20
                analysis['dpgf_compatibility']['detected_elements'].append(f"Colonnes de prix ({len(patterns['price_columns'])})")
            else:
                analysis['dpgf_compatibility']['missing_elements'].append("Aucune colonne de prix détectée")
            
            if patterns['unit_columns']:
                sheet_score += 15
                analysis['dpgf_compatibility']['detected_elements'].append(f"Colonnes d'unité ({len(patterns['unit_columns'])})")
            else:
                analysis['dpgf_compatibility']['missing_elements'].append("Aucune colonne d'unité détectée")
            
            if patterns['structure_analysis']['potential_element_rows']:
                element_count = len(patterns['structure_analysis']['potential_element_rows'])
                if element_count >= 10:
                    sheet_score += 20
                elif element_count >= 5:
                    sheet_score += 10
                else:
                    sheet_score += 5
                analysis['dpgf_compatibility']['detected_elements'].append(f"Éléments potentiels ({element_count})")
            else:
                analysis['dpgf_compatibility']['missing_elements'].append("Aucun élément détecté")
            
            total_score = max(total_score, sheet_score)
        
        analysis['dpgf_compatibility']['score'] = total_score / max_score
        
        # Générer des raisons pour le score
        if analysis['dpgf_compatibility']['score'] >= 0.8:
            analysis['dpgf_compatibility']['reasons'].append("Structure DPGF bien définie")
        elif analysis['dpgf_compatibility']['score'] >= 0.5:
            analysis['dpgf_compatibility']['reasons'].append("Structure DPGF partielle")
        else:
            analysis['dpgf_compatibility']['reasons'].append("Structure DPGF faible ou absente")
        
        return analysis
    
    def generate_recommendations(self, analysis: Dict) -> List[str]:
        """Génère des recommandations basées sur l'analyse"""
        recommendations = []
        
        score = analysis['dpgf_compatibility']['score']
        missing = analysis['dpgf_compatibility']['missing_elements']
        
        if score < 0.3:
            recommendations.append("🚨 Fichier probablement NON-DPGF - Vérifier l'identification automatique")
            recommendations.append("📋 Examiner manuellement le contenu du fichier")
        
        if "Aucun lot détecté" in missing:
            recommendations.append("🔍 Ajouter une ligne de lot explicite (ex: 'LOT 1 - Description')")
            recommendations.append("⚙️ Vérifier les mots-clés de détection des lots")
        
        if "Aucune section détectée" in missing:
            recommendations.append("📂 Structurer le fichier avec des sections/chapitres")
            recommendations.append("🏷️ Utiliser des mots-clés comme 'Chapitre', 'Section', 'Poste'")
        
        if "Aucune colonne de prix détectée" in missing:
            recommendations.append("💰 Ajouter une colonne 'Prix Unitaire' ou 'P.U.'")
            recommendations.append("📊 Vérifier que les prix sont dans des colonnes dédiées")
        
        if "Aucune colonne d'unité détectée" in missing:
            recommendations.append("📏 Ajouter une colonne 'Unité' avec les unités de mesure")
            recommendations.append("🔤 Utiliser des unités standard (m², ml, kg, ens, etc.)")
        
        if "Aucun élément détecté" in missing:
            recommendations.append("📝 Vérifier que chaque ligne d'élément a une désignation et un prix")
            recommendations.append("🔢 S'assurer que les valeurs numériques sont bien formatées")
        
        # Recommandations spécifiques selon le score
        if 0.3 <= score < 0.6:
            recommendations.append("🔧 Fichier partiellement compatible - Améliorer la structure")
            recommendations.append("📚 Consulter un exemple de DPGF type pour la structure")
        
        elif score >= 0.6:
            recommendations.append("✅ Structure DPGF détectée - Vérifier les détails de l'import")
            recommendations.append("🐛 Utiliser --debug-import pour voir les logs détaillés")
        
        return recommendations
    
    def save_analysis(self, output_file: str = None) -> str:
        """Sauvegarde l'analyse complète"""
        analysis = self.analyze_file_structure()
        recommendations = self.generate_recommendations(analysis)
        
        full_analysis = {
            'metadata': {
                'analyzed_at': datetime.now().isoformat(),
                'file_path': str(self.file_path),
                'analyzer_version': '1.0'
            },
            'analysis': analysis,
            'recommendations': recommendations
        }
        
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_name = "".join(c for c in self.file_path.stem if c.isalnum() or c in (' ', '-', '_'))[:30]
            output_file = f"analysis_{safe_name}_{timestamp}.json"
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(full_analysis, f, indent=2, ensure_ascii=False)
        
        return output_file
    
    def print_summary(self):
        """Affiche un résumé de l'analyse"""
        analysis = self.analyze_file_structure()
        recommendations = self.generate_recommendations(analysis)
        
        print("🔍 ANALYSE DÉTAILLÉE DU FICHIER DPGF")
        print("=" * 60)
        print(f"📄 Fichier: {analysis['file_info']['name']}")
        print(f"📊 Taille: {analysis['file_info']['size']:,} octets")
        print(f"📋 Feuilles: {analysis['file_info']['sheets_count']}")
        print()
        
        # Score de compatibilité
        score = analysis['dpgf_compatibility']['score']
        score_pct = score * 100
        
        if score >= 0.8:
            status = "🟢 EXCELLENTE"
        elif score >= 0.6:
            status = "🟡 BONNE"
        elif score >= 0.3:
            status = "🟠 FAIBLE"
        else:
            status = "🔴 TRÈS FAIBLE"
        
        print(f"🎯 COMPATIBILITÉ DPGF: {status} ({score_pct:.1f}%)")
        print()
        
        # Éléments détectés
        if analysis['dpgf_compatibility']['detected_elements']:
            print("✅ ÉLÉMENTS DÉTECTÉS:")
            for element in analysis['dpgf_compatibility']['detected_elements']:
                print(f"   • {element}")
            print()
        
        # Éléments manquants
        if analysis['dpgf_compatibility']['missing_elements']:
            print("❌ ÉLÉMENTS MANQUANTS:")
            for element in analysis['dpgf_compatibility']['missing_elements']:
                print(f"   • {element}")
            print()
        
        # Analyse par feuille
        print("📊 ANALYSE PAR FEUILLE:")
        for sheet_name, patterns in analysis['sheets_analysis'].items():
            print(f"   📋 {sheet_name}:")
            print(f"      • Lots: {len(patterns['lot_indicators'])}")
            print(f"      • Sections: {len(patterns['section_indicators'])}")
            print(f"      • Colonnes prix: {len(patterns['price_columns'])}")
            print(f"      • Colonnes unité: {len(patterns['unit_columns'])}")
            print(f"      • Éléments potentiels: {len(patterns['structure_analysis']['potential_element_rows'])}")
        print()
        
        # Recommandations
        if recommendations:
            print("💡 RECOMMANDATIONS:")
            for i, rec in enumerate(recommendations, 1):
                print(f"   {i}. {rec}")
            print()
        
        print("=" * 60)
    
    def close(self):
        """Ferme le fichier"""
        if self.workbook:
            self.workbook.close()

def main():
    """Fonction principale"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Analyse détaillée d'un fichier DPGF")
    parser.add_argument('file_path', help='Chemin vers le fichier DPGF à analyser')
    parser.add_argument('--save-analysis', action='store_true', 
                       help='Sauvegarder l\'analyse en JSON')
    parser.add_argument('--output-file', help='Fichier de sortie pour l\'analyse')
    
    args = parser.parse_args()
    
    try:
        analyzer = DPGFFileAnalyzer(args.file_path)
        
        if args.save_analysis:
            output_file = analyzer.save_analysis(args.output_file)
            print(f"📄 Analyse sauvegardée: {output_file}")
        else:
            analyzer.print_summary()
    
    except Exception as e:
        print(f"❌ Erreur: {e}")
        return 1
    
    finally:
        if 'analyzer' in locals():
            analyzer.close()
    
    return 0

if __name__ == "__main__":
    sys.exit(main())
