"""
Module d'aide pour l'import des fichiers DPGF provenant de SharePoint.
Ce module contient des fonctions spécialisées pour traiter les formats spécifiques
des fichiers Excel générés ou modifiés via SharePoint.
"""

import re
import os
import pandas as pd
import logging
from typing import List, Dict, Optional, Tuple, Any
from pathlib import Path
import openpyxl

# Configuration du logger
logger = logging.getLogger(__name__)

class SharePointExcelHelper:
    """
    Classe d'aide pour traiter les fichiers Excel SharePoint.
    Contient des méthodes spécifiques pour détecter les feuilles pertinentes,
    les en-têtes et les sections dans un format typique de SharePoint.
    """
    
    def __init__(self, file_path: str):
        """
        Initialise le helper avec un chemin de fichier Excel
        
        Args:
            file_path: Chemin complet vers le fichier Excel
        """
        self.file_path = file_path
        self.workbook = None
        self.sheet_names = []
        self.selected_sheet = None
        self.df = None
        
        # Charger le classeur
        self._load_workbook()
    
    def _load_workbook(self):
        """Charge le classeur Excel et liste ses feuilles"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            logger.info(f"Fichier Excel chargé: {self.file_path}")
            logger.info(f"Feuilles disponibles: {', '.join(self.sheet_names)}")
        except Exception as e:
            logger.error(f"Erreur lors du chargement du classeur: {e}")
            raise
    
    def is_sharepoint_format(self) -> bool:
        """
        Détecte si le fichier semble être au format SharePoint
        
        Returns:
            True si le format SharePoint est détecté, False sinon
        """
        # Indices de format SharePoint: plusieurs feuilles, noms spécifiques, etc.
        if len(self.sheet_names) >= 3:  # Souvent 3+ feuilles (Infos, Page de garde, LOT xx)
            return True
            
        # Vérifier les noms spécifiques souvent présents dans les fichiers SharePoint
        sharepoint_sheet_patterns = ['infos', 'page de garde', 'lot', 'récap', 'sommaire']
        if any(any(pattern in name.lower() for pattern in sharepoint_sheet_patterns) 
               for name in self.sheet_names):
            return True
            
        # Si le fichier contient une feuille "LOT xx", c'est souvent un indicateur
        lot_sheets = [name for name in self.sheet_names 
                     if re.search(r'lot\s*\d+', name.lower())]
        if lot_sheets:
            return True
            
        return False
    
    def select_best_sheet(self) -> str:
        """
        Sélectionne la meilleure feuille à utiliser pour l'import DPGF
        
        Returns:
            Nom de la feuille sélectionnée
        """
        # Scores pour évaluer la pertinence de chaque feuille
        sheet_scores = {}
        
        # Évaluer chaque feuille
        for sheet_name in self.sheet_names:
            score = 0
            lower_name = sheet_name.lower()
            
            # 1. Vérifier si le nom contient "lot" (fort indicateur)
            if 'lot' in lower_name:
                score += 5
                # Bonus si suivi d'un numéro (ex: "LOT 04")
                if re.search(r'lot\s*\d+', lower_name):
                    score += 3
            
            # 2. Vérifier la présence d'autres mots-clés pertinents
            keywords = ['dpgf', 'prix', 'quantitatif', 'detail', 'détail', 'chiffrage']
            for keyword in keywords:
                if keyword in lower_name:
                    score += 2
            
            # 3. Éléments négatifs (feuilles généralement non pertinentes)
            negative_keywords = ['info', 'garde', 'sommaire', 'récap']
            for neg_keyword in negative_keywords:
                if neg_keyword in lower_name:
                    score -= 3
            
            # 4. Charger un aperçu de la feuille pour vérifier si elle contient des données pertinentes
            try:
                # Lire les 10 premières lignes
                df_preview = pd.read_excel(self.file_path, sheet_name=sheet_name, 
                                         nrows=10, engine='openpyxl')
                
                # Vérifier le nombre de colonnes et de données
                score += min(len(df_preview.columns), 5)  # +1 par colonne jusqu'à 5
                
                # Chercher des mots clés typiques des tableaux DPGF
                sample_text = ' '.join(str(val).lower() for val in df_preview.values.flatten() if pd.notna(val))
                dpgf_keywords = ['désignation', 'unité', 'quantité', 'prix', 'total', 'ht', 'ttc']
                for keyword in dpgf_keywords:
                    if keyword in sample_text:
                        score += 1
                
                # Vérifier la présence de numéros de sections (5.1, 5.1.1)
                if re.search(r'\d+\.\d+(?:\.\d+)?', sample_text):
                    score += 3
                    
                # Nombre de cellules contenant des données
                non_empty_cells = df_preview.count().sum()
                score += min(non_empty_cells // 10, 5)  # +1 pour chaque 10 cellules non vides (max 5)
                
            except Exception as e:
                logger.warning(f"Erreur lors de l'analyse de la feuille {sheet_name}: {e}")
                score -= 2  # Pénalité pour les erreurs
            
            sheet_scores[sheet_name] = score
            logger.info(f"Feuille '{sheet_name}': score {score}")
        
        # Sélectionner la feuille avec le meilleur score
        if not sheet_scores:
            # Si pas de scores (erreurs), prendre la première feuille
            selected_sheet = self.sheet_names[0]
        else:
            selected_sheet = max(sheet_scores.items(), key=lambda x: x[1])[0]
        
        self.selected_sheet = selected_sheet
        logger.info(f"Feuille sélectionnée: '{selected_sheet}' (score: {sheet_scores.get(selected_sheet, 'N/A')})")
        
        return selected_sheet
    
    def load_selected_sheet(self) -> pd.DataFrame:
        """
        Charge la feuille sélectionnée dans un DataFrame
        
        Returns:
            DataFrame contenant les données de la feuille
        """
        if not self.selected_sheet:
            self.select_best_sheet()
            
        try:
            self.df = pd.read_excel(self.file_path, sheet_name=self.selected_sheet, engine='openpyxl')
            logger.info(f"Feuille '{self.selected_sheet}' chargée: {self.df.shape[0]} lignes, {self.df.shape[1]} colonnes")
            return self.df
        except Exception as e:
            logger.error(f"Erreur lors du chargement de la feuille '{self.selected_sheet}': {e}")
            raise
    
    def find_header_row_sharepoint(self) -> Optional[int]:
        """
        Trouve la ligne d'en-tête dans un format SharePoint
        qui contient souvent Désignation/Quantité/Prix unitaire/Prix total
        
        Returns:
            Index de la ligne d'en-tête ou None si non trouvée
        """
        if self.df is None:
            self.load_selected_sheet()
        
        # Patterns pour reconnaître les en-têtes dans les fichiers SharePoint
        header_patterns = {
            'designation': [r'désignation', r'designation', r'libellé', r'libelle', r'description', 
                          r'des ouvrages', r'intitulé', r'detail', r'détail', r'ouvrages'],
            'unite': [r'unité', r'unite', r'u\.?$', r'un\.?$', r'un$'],
            'quantite': [r'quantité', r'quantite', r'qté\.?', r'qt\.?', r'quant\.?', r'qté', r'qte'],
            'prix_unitaire': [r'prix\s*(?:unitaire|unit\.?)', r'p\.u\.', r'pu', r'pu\s*ht'],
            'prix_total': [r'prix\s*(?:total|tot\.?)', r'montant', r'p\.t\.', r'pt', r'total', r'ttc']
        }
        
        # Parcourir les 20 premières lignes (fichiers SharePoint ont souvent des en-têtes tardifs)
        for i in range(min(20, len(self.df))):
            # Convertir la ligne en texte pour recherche plus facile
            row_values = [str(val).strip().lower() if pd.notna(val) else "" for val in self.df.iloc[i].values]
            row_text = " ".join(row_values)
            
            # Compter combien de types d'en-têtes sont présents
            matches = 0
            for col_type, patterns in header_patterns.items():
                if any(re.search(pattern, row_text, re.IGNORECASE) for pattern in patterns):
                    matches += 1
                else:
                    # Chercher dans chaque cellule individuellement
                    for cell_text in row_values:
                        if any(re.search(f"^{pattern}$", cell_text, re.IGNORECASE) for pattern in patterns):
                            matches += 1
                            break
            
            # Si la ligne contient au moins 2 types d'en-têtes, c'est probablement la bonne
            if matches >= 2:
                logger.info(f"Ligne d'en-tête SharePoint trouvée (ligne {i+1}): {matches} types d'en-têtes")
                return i
        
        logger.warning("Aucune ligne d'en-tête SharePoint trouvée")
        return None
    
    def detect_column_indices_sharepoint(self, header_row_idx: Optional[int] = None) -> Dict[str, Optional[int]]:
        """
        Détermine l'indice des colonnes importantes en se basant sur l'en-tête
        pour les fichiers SharePoint qui ont souvent des structures différentes
        
        Args:
            header_row_idx: Indice de la ligne d'en-tête
            
        Returns:
            Dictionnaire avec les indices des colonnes
        """
        if self.df is None:
            self.load_selected_sheet()
            
        if header_row_idx is None:
            header_row_idx = self.find_header_row_sharepoint()
            
        # Initialiser les indices à None
        column_indices = {
            'designation': None,
            'unite': None,
            'quantite': None,
            'prix_unitaire': None,
            'prix_total': None
        }
        
        # Si pas d'en-tête trouvé, essayer de détecter par le contenu
        if header_row_idx is None:
            # Pour SharePoint, chercher les colonnes qui contiennent des numéros de section (5.1, 5.1.1)
            for col_idx in range(min(10, len(self.df.columns))):
                section_pattern_count = 0
                description_text_count = 0
                
                # Examiner les 30 premières lignes
                for row_idx in range(min(30, len(self.df))):
                    cell_value = self.df.iloc[row_idx, col_idx]
                    if pd.isna(cell_value):
                        continue
                        
                    cell_str = str(cell_value).strip()
                    
                    # Compter les patterns de numéros de section
                    if re.match(r'^\d+\.\d+(?:\.\d+)?$', cell_str):
                        section_pattern_count += 1
                        
                    # Compter les cellules avec du texte descriptif
                    elif len(cell_str) > 15:  # Texte substantiel
                        description_text_count += 1
                
                # Si cette colonne contient des numéros de section, c'est probablement celle juste avant la désignation
                if section_pattern_count > 3:
                    # La colonne de désignation est généralement la suivante
                    if col_idx + 1 < len(self.df.columns):
                        column_indices['designation'] = col_idx + 1
                        logger.info(f"Colonne de désignation SharePoint détectée par numéros de section: {col_idx + 1}")
                        break
                        
                # Si cette colonne contient beaucoup de texte descriptif, c'est probablement la désignation
                elif description_text_count > 5:
                    column_indices['designation'] = col_idx
                    logger.info(f"Colonne de désignation SharePoint détectée par texte: {col_idx}")
                    break
            
            # Chercher les colonnes numériques qui pourraient être quantité/prix
            numeric_cols = []
            for col_idx in range(min(15, len(self.df.columns))):
                numeric_count = 0
                for row_idx in range(min(30, len(self.df))):
                    cell_value = self.df.iloc[row_idx, col_idx]
                    if pd.isna(cell_value):
                        continue
                        
                    # Vérifier si la valeur est numérique
                    try:
                        float(str(cell_value).replace(',', '.'))
                        numeric_count += 1
                    except (ValueError, TypeError):
                        pass
                        
                if numeric_count > 5:
                    numeric_cols.append((col_idx, numeric_count))
            
            # Trier les colonnes numériques par nombre de valeurs
            numeric_cols.sort(key=lambda x: x[1], reverse=True)
            
            # Assigner les colonnes numériques par ordre logique
            if len(numeric_cols) >= 3:
                column_indices['quantite'] = numeric_cols[0][0]
                column_indices['prix_unitaire'] = numeric_cols[1][0]
                column_indices['prix_total'] = numeric_cols[2][0]
                logger.info(f"Colonnes numériques SharePoint détectées: quantité={numeric_cols[0][0]}, "
                         f"prix unitaire={numeric_cols[1][0]}, prix total={numeric_cols[2][0]}")
            elif len(numeric_cols) == 2:
                column_indices['prix_unitaire'] = numeric_cols[0][0]
                column_indices['prix_total'] = numeric_cols[1][0]
                logger.info(f"Colonnes de prix SharePoint détectées: prix unitaire={numeric_cols[0][0]}, "
                         f"prix total={numeric_cols[1][0]}")
            
        else:
            # Si un en-tête a été trouvé, utiliser la détection standard mais adaptée à SharePoint
            header_row = [str(val).strip().lower() if pd.notna(val) else "" for val in self.df.iloc[header_row_idx].values]
            
            # Patterns pour chaque type de colonne
            patterns = {
                'designation': [r'désignation', r'designation', r'libellé', r'libelle', r'description', 
                              r'des ouvrages', r'intitulé', r'detail', r'détail'],
                'unite': [r'unité', r'unite', r'u\.?$', r'un\.?$', r'un$'],
                'quantite': [r'quantité', r'quantite', r'qté\.?', r'qt\.?', r'quant\.?', r'qté', r'qte'],
                'prix_unitaire': [r'prix\s*(?:unitaire|unit\.?)', r'p\.u\.', r'pu', r'pu\s*ht'],
                'prix_total': [r'prix\s*(?:total|tot\.?)', r'montant', r'p\.t\.', r'pt', r'total', r'ttc']
            }
            
            # Chercher chaque pattern dans l'en-tête
            for col_name, col_patterns in patterns.items():
                for col_idx, cell_text in enumerate(header_row):
                    cell_text = cell_text.lower()
                    for pattern in col_patterns:
                        if re.search(pattern, cell_text, re.IGNORECASE):
                            column_indices[col_name] = col_idx
                            logger.info(f"Colonne SharePoint '{col_name}' détectée: indice {col_idx}, valeur: '{header_row[col_idx]}'")
                            break
                    if column_indices[col_name] is not None:
                        break
        
        # Vérifications supplémentaires spécifiques à SharePoint
        # Si on n'a pas trouvé la colonne de désignation, chercher par position après les numéros de section
        if column_indices['designation'] is None:
            # Dans les fichiers SharePoint, la désignation est souvent la 2e colonne après le numéro
            # Chercher une colonne avec beaucoup de texte
            max_text_col = 0
            max_text_length = 0
            
            for col_idx in range(min(5, len(self.df.columns))):
                total_text_length = 0
                for row_idx in range(min(header_row_idx + 1 if header_row_idx else 0, 30)):
                    if row_idx < len(self.df) and col_idx < len(self.df.columns):
                        cell_value = self.df.iloc[row_idx, col_idx]
                        if pd.notna(cell_value):
                            total_text_length += len(str(cell_value))
                
                if total_text_length > max_text_length:
                    max_text_length = total_text_length
                    max_text_col = col_idx
            
            column_indices['designation'] = max_text_col
            logger.info(f"Colonne de désignation SharePoint détectée par analyse de contenu: {max_text_col}")
        
        return column_indices
    
    def detect_sections_sharepoint(self, header_row: Optional[int] = None) -> List[Dict]:
        """
        Détecte les sections dans un format SharePoint en utilisant la numérotation hiérarchique (5.1, 5.1.1)
        
        Args:
            header_row: Index de la ligne d'en-tête
            
        Returns:
            Liste de dictionnaires représentant les sections
        """
        if self.df is None:
            self.load_selected_sheet()
            
        if header_row is None:
            header_row = self.find_header_row_sharepoint()
        
        # Démarrer après la ligne d'en-tête
        start_row = header_row + 1 if header_row is not None else 0
        
        # Patterns spécifiques pour SharePoint
        section_patterns = {
            'numbered_sp': re.compile(r'^(\d+(?:\.\d+)+)\s+(.*)'),   # Format SharePoint: 5.1, 5.1.1, etc.
            'numbered': re.compile(r'^(\d+(?:\.\d+)*)\s+(.*)'),      # Formats numérotés génériques
            'uppercase': re.compile(r'^([A-Z][A-Z\s\d\.]+)$')        # Titres en majuscules
        }
        
        sections = []
        current_level = {}  # Garde trace du dernier numéro de section par niveau
        
        # Parcourir toutes les lignes à partir de la ligne d'en-tête
        for i in range(start_row, len(self.df)):
            row_values = self.df.iloc[i].values
            
            # Obtenir le texte de la première cellule qui pourrait contenir un numéro de section
            section_cell = None
            designation_cell = None
            
            # Chercher dans les 3 premières colonnes
            for col_idx in range(min(3, len(row_values))):
                if pd.notna(row_values[col_idx]):
                    cell_text = str(row_values[col_idx]).strip()
                    
                    # Si cette cellule contient un potentiel numéro de section
                    if re.match(r'^\d+\.\d+', cell_text):
                        section_cell = cell_text
                        
                        # La désignation peut être dans la même cellule ou dans la suivante
                        if " " in cell_text and len(cell_text.split(" ", 1)[1]) > 3:
                            # Le numéro et la désignation sont dans la même cellule
                            designation_cell = cell_text.split(" ", 1)[1]
                        elif col_idx + 1 < len(row_values) and pd.notna(row_values[col_idx + 1]):
                            # La désignation est dans la cellule suivante
                            designation_cell = str(row_values[col_idx + 1]).strip()
                        break
            
            # Si on n'a pas trouvé de numéro de section, vérifier si la ligne entière est une section
            if section_cell is None:
                # Joindre les valeurs non-nulles de la ligne pour analyse
                row_text = " ".join([str(val).strip() for val in row_values if pd.notna(val)])
                
                for pattern_name, pattern in section_patterns.items():
                    match = pattern.match(row_text)
                    if match:
                        # Format dépend du pattern
                        if pattern_name == 'uppercase':
                            section_num = f"S{hash(match.group(1)) % 10000:04d}"  # Générer un ID unique
                            section_title = match.group(1)
                            hierarchical_level = 1  # Les sections en majuscules sont des titres principaux
                        else:
                            section_num = match.group(1)
                            section_title = match.group(2) if len(match.groups()) > 1 else ""
                            # Calculer le niveau hiérarchique
                            hierarchical_level = section_num.count('.') + 1
                        
                        sections.append({
                            'row': i,
                            'type': 'section',
                            'data': {
                                'numero_section': section_num,
                                'titre_section': section_title,
                                'niveau_hierarchique': hierarchical_level
                            }
                        })
                        
                        current_level[hierarchical_level] = section_num
                        logger.info(f"Section détectée (ligne {i+1}): {section_num} - {section_title} (niveau {hierarchical_level})")
                        break
                        
            # Si on a trouvé un numéro de section spécifique SharePoint
            elif section_cell and designation_cell:
                # Extraire le numéro de section (5.1, 5.1.1, etc.)
                match = re.match(r'^(\d+\.\d+(?:\.\d+)*)', section_cell)
                if match:
                    section_num = match.group(1)
                    section_title = designation_cell
                    hierarchical_level = section_num.count('.') + 1
                    
                    sections.append({
                        'row': i,
                        'type': 'section',
                        'data': {
                            'numero_section': section_num,
                            'titre_section': section_title,
                            'niveau_hierarchique': hierarchical_level
                        }
                    })
                    
                    current_level[hierarchical_level] = section_num
                    logger.info(f"Section SharePoint détectée (ligne {i+1}): {section_num} - {section_title} (niveau {hierarchical_level})")
        
        return sections
    
    def analyze_sharepoint_file(self) -> Dict:
        """
        Analyse complète d'un fichier SharePoint avec détection des feuilles, en-têtes et sections
        
        Returns:
            Dictionnaire avec les résultats de l'analyse
        """
        results = {
            'file_path': self.file_path,
            'is_sharepoint': self.is_sharepoint_format(),
            'sheet_analysis': {},
            'header_row': None,
            'columns': {},
            'sections': []
        }
        
        if results['is_sharepoint']:
            # Sélectionner et charger la meilleure feuille
            selected_sheet = self.select_best_sheet()
            self.load_selected_sheet()
            results['sheet_analysis'] = {
                'selected_sheet': selected_sheet,
                'all_sheets': self.sheet_names
            }
            
            # Trouver l'en-tête
            header_row = self.find_header_row_sharepoint()
            results['header_row'] = header_row
            
            # Détecter les colonnes
            columns = self.detect_column_indices_sharepoint(header_row)
            results['columns'] = columns
            
            # Détecter les sections
            sections = self.detect_sections_sharepoint(header_row)
            results['sections'] = sections
            
        return results


def is_sharepoint_file(file_path: str) -> bool:
    """
    Détecte si un fichier Excel semble être au format SharePoint
    
    Args:
        file_path: Chemin vers le fichier Excel à analyser
        
    Returns:
        True si le fichier semble être au format SharePoint, False sinon
    """
    try:
        helper = SharePointExcelHelper(file_path)
        return helper.is_sharepoint_format()
    except Exception as e:
        logger.error(f"Erreur lors de la détection du format SharePoint: {e}")
        return False


def analyze_sharepoint_file(file_path: str) -> Dict:
    """
    Analyse un fichier Excel SharePoint et retourne les informations pertinentes
    
    Args:
        file_path: Chemin vers le fichier Excel à analyser
        
    Returns:
        Dictionnaire avec les résultats de l'analyse
    """
    helper = SharePointExcelHelper(file_path)
    return helper.analyze_sharepoint_file()
