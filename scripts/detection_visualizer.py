"""
Visualiseur de détection DPGF - Outil de diagnostic visuel pour l'import DPGF
Ce script permet de visualiser les résultats de détection des lots et sections,
et d'explorer les problèmes potentiels sur un fichier Excel.
"""

import os
import sys
import json
import argparse
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import to_rgba
import re
from typing import List, Dict, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import Color

# Ajouter le répertoire principal au path pour importer les modules
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from scripts.import_complete import ExcelParser, ColumnMapping, ErrorReporter
from scripts.enhanced_logging import get_import_logger


class DPGFViewer:
    """
    Visualiseur pour diagnostiquer les problèmes de détection DPGF
    """
    
    def __init__(self, file_path: str, output_dir: str = "visualizations"):
        """
        Initialise le visualiseur
        
        Args:
            file_path: Chemin vers le fichier DPGF à analyser
            output_dir: Répertoire de sortie pour les visualisations
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Fichier non trouvé: {self.file_path}")
        
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        self.column_mapper = ColumnMapping()
        self.error_reporter = ErrorReporter()
        self.logger = get_import_logger(file_path)
        
        # Charger le fichier avec ExcelParser
        self.parser = ExcelParser(str(self.file_path), self.column_mapper, self.error_reporter, self.logger)
        
        # Résultats de l'analyse
        self.header_row = None
        self.lots = []
        self.items = []
        self.sections = []
        self.elements = []
        self.column_indices = {}
        self.success = False
        
        # Réaliser l'analyse
        self._analyze_file()
    
    def _analyze_file(self):
        """
        Réalise l'analyse complète du fichier DPGF
        """
        try:
            # 1. Détecter les lots
            self.lots = self.parser.find_lot_headers()
            
            # 2. Trouver la ligne d'en-tête
            self.header_row = self.parser.find_header_row()
            
            # 3. Détecter les colonnes
            if not self.parser.headers_detected and self.header_row is not None:
                self.parser.detect_column_indices(self.header_row)
            
            self.column_indices = {
                "designation": self.parser.col_designation,
                "unite": self.parser.col_unite,
                "quantite": self.parser.col_quantite,
                "prix_unitaire": self.parser.col_prix_unitaire,
                "prix_total": self.parser.col_prix_total
            }
            
            # 4. Détecter les sections et éléments
            self.items = self.parser.detect_sections_and_elements(self.header_row)
            
            # Séparer les sections et éléments
            self.sections = [item for item in self.items if item['type'] == 'section']
            self.elements = [item for item in self.items if item['type'] == 'element']
            
            self.success = True
        except Exception as e:
            print(f"⚠️ Erreur lors de l'analyse: {e}")
            self.success = False
    
    def generate_html_report(self):
        """
        Génère un rapport HTML interactif pour visualiser les détections
        """
        if not self.success:
            print("⚠️ L'analyse a échoué, impossible de générer le rapport HTML")
            return
        
        output_file = self.output_dir / f"{self.file_path.stem}_report.html"
        
        # Préparer les données pour le tableau
        df = self.parser.df.copy()
        
        # Marquer les types de lignes
        row_types = [""] * len(df)
        
        # Marquer la ligne d'en-tête
        if self.header_row is not None:
            row_types[self.header_row] = "header"
        
        # Marquer les sections
        section_rows = [item['row'] for item in self.sections]
        for row in section_rows:
            if 0 <= row < len(row_types):
                row_types[row] = "section"
        
        # Marquer les éléments
        element_rows = [item['row'] for item in self.elements]
        for row in element_rows:
            if 0 <= row < len(row_types):
                row_types[row] = "element"
        
        # Autres informations importantes
        lot_info = ""
        if self.lots:
            lot_info = f"<strong>Lots détectés:</strong> {', '.join([f'Lot {num} - {name}' for num, name in self.lots])}<br>"
        else:
            lot_info = "<strong>Aucun lot détecté</strong><br>"
        
        # Informations sur les colonnes
        col_info = ""
        for name, idx in self.column_indices.items():
            if idx is not None:
                col_name = chr(65 + idx) if idx < 26 else chr(64 + idx // 26) + chr(65 + idx % 26)
                col_info += f"<strong>{name}</strong>: colonne {col_name}, "
        
        # Mettre à jour les données
        df['__row_type'] = row_types
        df['__row_idx'] = range(len(df))
        
        # Générer le tableau HTML avec styles
        html_data = df.to_html(index=False)
        
        # Ajouter les styles CSS pour les types de lignes
        css = """
        <style>
            table { border-collapse: collapse; width: 100%; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            tr.header { background-color: #f2dede; font-weight: bold; }
            tr.section { background-color: #d9edf7; font-weight: bold; }
            tr.element { background-color: #dff0d8; }
            th { background-color: #4CAF50; color: white; }
            .summary { background-color: #f8f9fa; padding: 15px; margin-bottom: 20px; border: 1px solid #ddd; }
            h2 { color: #2c3e50; }
        </style>
        <script>
            function highlightRow(rowIdx) {
                const rows = document.querySelectorAll('tr');
                rows.forEach(row => {
                    if(row.getAttribute('data-row-idx') === rowIdx) {
                        row.style.backgroundColor = '#ffffcc';
                    }
                });
            }
        </script>
        """
        
        # Créer le document HTML complet
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Rapport de détection DPGF - {self.file_path.name}</title>
            {css}
        </head>
        <body>
            <h2>Rapport de détection DPGF - {self.file_path.name}</h2>
            <div class="summary">
                {lot_info}
                <strong>Ligne d'en-tête:</strong> {self.header_row + 1 if self.header_row is not None else 'Non détectée'}<br>
                <strong>Sections détectées:</strong> {len(self.sections)}<br>
                <strong>Éléments détectés:</strong> {len(self.elements)}<br>
                <strong>Colonnes:</strong> {col_info}<br>
            </div>
            
            <h3>Contenu du fichier</h3>
        """
        
        # Insérer le tableau avec coloration des lignes selon leur type
        table_html = html_data.replace('<table', '<table id="dpgf-table"')
        
        # Ajouter des classes aux lignes selon leur type
        for row_type in ["header", "section", "element"]:
            table_html = table_html.replace(f'<td>{row_type}</td>', f'<td>{row_type}</td>', -1)
        
        # Ajouter les attributs data-row-idx pour le highlight
        for i in range(len(df)):
            table_html = table_html.replace(f'<td>{i}</td>', f'<td>{i}</td>', 1)
        
        html_content += table_html
        html_content += """
        <script>
            // Ajouter les classes aux lignes selon leur type
            document.querySelectorAll('#dpgf-table tbody tr').forEach(row => {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 1) {
                    const rowType = cells[cells.length-2].innerText;
                    const rowIdx = cells[cells.length-1].innerText;
                    if (rowType) {
                        row.classList.add(rowType);
                    }
                    row.setAttribute('data-row-idx', rowIdx);
                }
            });
        </script>
        </body>
        </html>
        """
        
        # Écrire le fichier HTML
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"✓ Rapport HTML généré: {output_file}")
        return output_file
    
    def create_annotated_excel(self):
        """
        Crée une version annotée du fichier Excel avec les détections
        """
        if not self.success:
            print("⚠️ L'analyse a échoué, impossible de créer le fichier Excel annoté")
            return
        
        output_file = self.output_dir / f"{self.file_path.stem}_annotated.xlsx"
        
        # Copier le fichier original
        try:
            # Charger le workbook
            wb = load_workbook(self.file_path)
            sheet = wb.active
            
            # Définir les styles
            header_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            section_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
            element_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            bold_font = Font(bold=True)
            
            # Appliquer les styles
            
            # Ligne d'en-tête
            if self.header_row is not None:
                for cell in sheet[self.header_row + 1]:
                    cell.fill = header_fill
                    cell.font = bold_font
            
            # Sections
            for section in self.sections:
                row = section['row'] + 1  # Excel rows are 1-indexed
                for cell in sheet[row]:
                    cell.fill = section_fill
                    cell.font = bold_font
            
            # Éléments
            for element in self.elements:
                row = element['row'] + 1
                for cell in sheet[row]:
                    cell.fill = element_fill
            
            # Sauvegarder le fichier annoté
            wb.save(output_file)
            
            print(f"✓ Fichier Excel annoté généré: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"⚠️ Erreur lors de la création du fichier Excel annoté: {e}")
            return None
    
    def plot_detection_heatmap(self):
        """
        Génère une heatmap des détections pour visualiser facilement les problèmes
        """
        if not self.success:
            print("⚠️ L'analyse a échoué, impossible de générer la heatmap")
            return
        
        output_file = self.output_dir / f"{self.file_path.stem}_heatmap.png"
        
        try:
            # Créer une matrice pour la heatmap
            df = self.parser.df
            nrows, ncols = df.shape
            
            # Limiter à 100 lignes et 15 colonnes pour la lisibilité
            max_rows = min(nrows, 100)
            max_cols = min(ncols, 15)
            
            # Créer une matrice de valeurs pour la heatmap
            heatmap_data = np.zeros((max_rows, max_cols))
            
            # Marquer les types de cellules
            # 1: ligne d'en-tête, 2: section, 3: élément, 0: autre
            
            # Ligne d'en-tête
            if self.header_row is not None and self.header_row < max_rows:
                heatmap_data[self.header_row, :] = 1
            
            # Sections
            for section in self.sections:
                row = section['row']
                if row < max_rows:
                    heatmap_data[row, :] = 2
            
            # Éléments
            for element in self.elements:
                row = element['row']
                if row < max_rows:
                    heatmap_data[row, :] = 3
            
            # Créer la figure
            plt.figure(figsize=(12, 20))
            plt.pcolormesh(heatmap_data, cmap=plt.cm.get_cmap('RdYlBu', 4), edgecolors='gray', linewidth=0.01)
            
            # Ajouter une colorbar
            cbar = plt.colorbar(ticks=[0.5, 1.5, 2.5, 3.5])
            cbar.set_ticklabels(['Autre', 'En-tête', 'Section', 'Élément'])
            
            # Définir les limites des axes
            plt.xlim(0, max_cols)
            plt.ylim(0, max_rows)
            
            # Inverser l'axe Y pour que la première ligne soit en haut
            plt.gca().invert_yaxis()
            
            # Ajouter les labels des axes
            plt.xlabel('Colonnes')
            plt.ylabel('Lignes')
            plt.title(f'Heatmap des détections - {self.file_path.name}')
            
            # Sauvegarder la figure
            plt.tight_layout()
            plt.savefig(output_file, dpi=100)
            plt.close()
            
            print(f"✓ Heatmap générée: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"⚠️ Erreur lors de la génération de la heatmap: {e}")
            return None


def main():
    """Point d'entrée principal"""
    parser = argparse.ArgumentParser(description="Visualiseur de détection DPGF")
    parser.add_argument('--file', '-f', type=str, required=True,
                       help="Chemin du fichier DPGF à analyser")
    parser.add_argument('--output-dir', '-o', type=str, default="visualizations",
                       help="Répertoire de sortie pour les visualisations")
    
    args = parser.parse_args()
    
    try:
        print(f"📊 Analyse du fichier: {args.file}")
        viewer = DPGFViewer(args.file, args.output_dir)
        
        # Générer les visualisations
        html_report = viewer.generate_html_report()
        excel_file = viewer.create_annotated_excel()
        heatmap = viewer.plot_detection_heatmap()
        
        print("\n✓ Visualisations générées avec succès!")
        print(f"📄 Rapport HTML: {html_report}")
        print(f"📊 Fichier Excel annoté: {excel_file}")
        print(f"🔥 Heatmap: {heatmap}")
    
    except Exception as e:
        print(f"⚠️ Erreur: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
